"""
Build docs/food-programme-rp-effort-calculator.xlsx

Companion artefact to docs/food-programme-rp-sop.md. Sizes the RP team
required to staff the field SOP at any scale of the food programme.

Run: python3 scripts/build-rp-effort-calculator.py
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

OUT = "docs/food-programme-rp-effort-calculator.xlsx"

# ── styling ────────────────────────────────────────────────────────────────
TITLE = Font(name="Calibri", size=16, bold=True, color="1F2937")
H1    = Font(name="Calibri", size=12, bold=True, color="111827")
H2    = Font(name="Calibri", size=11, bold=True, color="111827")
BODY  = Font(name="Calibri", size=11, color="111827")
HINT  = Font(name="Calibri", size=10, italic=True, color="6B7280")
INPUT_FONT  = Font(name="Calibri", size=11, bold=True, color="0B3D2E")
OUTPUT_FONT = Font(name="Calibri", size=11, bold=True, color="1E3A8A")
TOTAL_FONT  = Font(name="Calibri", size=12, bold=True, color="111827")

INPUT_FILL  = PatternFill("solid", fgColor="ECFDF5")  # mint  — editable
OUTPUT_FILL = PatternFill("solid", fgColor="EFF6FF")  # blue  — computed
TOTAL_FILL  = PatternFill("solid", fgColor="FEF3C7")  # amber — headline
SECTION_FILL = PatternFill("solid", fgColor="F3F4F6")

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_cell(cell, *, font=None, fill=None, align=None, border=BORDER, fmt=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = border
    if fmt: cell.number_format = fmt


LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

FMT_INT     = "#,##0"
FMT_INR     = "₹#,##0"
FMT_INR_DEC = "₹#,##0.00"
FMT_PCT     = "0%"


# ── workbook ───────────────────────────────────────────────────────────────
wb = Workbook()

# ===========================================================================
# Sheet 1 — Calculator
# ===========================================================================
ws = wb.active
ws.title = "Calculator"

ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 44
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 52

row = 1
ws.cell(row=row, column=2, value="RP Effort Calculator — Food Programme").font = TITLE
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
row += 1
c = ws.cell(row=row, column=2,
            value="Edit only mint-green cells. Blue cells are computed. See Assumptions sheet for logic.")
c.font = HINT
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
row += 2

# ── Section: Programme scale (inputs) ──────────────────────────────────────
def section_header(row, title):
    c = ws.cell(row=row, column=2, value=title)
    c.font = H1
    c.fill = SECTION_FILL
    c.alignment = LEFT
    for col in (3, 4, 5):
        ws.cell(row=row, column=col).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    return row + 1


def input_row(row, label, value, unit, note, name=None, fmt=FMT_INT):
    style_cell(ws.cell(row=row, column=2, value=label), font=BODY, align=LEFT)
    vc = ws.cell(row=row, column=3, value=value)
    style_cell(vc, font=INPUT_FONT, fill=INPUT_FILL, align=RIGHT, fmt=fmt)
    style_cell(ws.cell(row=row, column=4, value=unit), font=HINT, align=CENTER)
    style_cell(ws.cell(row=row, column=5, value=note), font=HINT, align=LEFT)
    if name:
        ref = f"Calculator!${get_column_letter(3)}${row}"
        wb.defined_names[name] = DefinedName(name=name, attr_text=ref)
    return row + 1


def output_row(row, label, formula, unit, note, name=None, fmt=FMT_INT, headline=False):
    style_cell(ws.cell(row=row, column=2, value=label),
               font=TOTAL_FONT if headline else BODY, align=LEFT)
    vc = ws.cell(row=row, column=3, value=formula)
    style_cell(
        vc,
        font=TOTAL_FONT if headline else OUTPUT_FONT,
        fill=TOTAL_FILL if headline else OUTPUT_FILL,
        align=RIGHT, fmt=fmt,
    )
    style_cell(ws.cell(row=row, column=4, value=unit), font=HINT, align=CENTER)
    style_cell(ws.cell(row=row, column=5, value=note), font=HINT, align=LEFT)
    if name:
        ref = f"Calculator!${get_column_letter(3)}${row}"
        wb.defined_names[name] = DefinedName(name=name, attr_text=ref)
    return row + 1


row = section_header(row, "1. Programme scale (inputs)")
row = input_row(row, "Total meals served per day",                1500, "meals",   "Launch: 1,500. Mid-scale: 5,000. Full: ~20,000.",                 "meals_per_day")
row = input_row(row, "Number of kitchens",                           1, "kitchens","One per geographical cluster. Cannot parallel-audit two from one RP.", "num_kitchens")
row = input_row(row, "Meals per truck per day",                    750, "meals",   "TATA Ace × 6 containers × ~125 portions. Edit if fleet changes.",      "meals_per_truck")
row = input_row(row, "Meals per DP per day (avg)",                 300, "meals",   "Sampark default 300; CFAR mix is 150–200. Use weighted avg.",          "meals_per_dp")
row = input_row(row, "DP visits per DP per week (RP coverage rule)", 2, "visits",  "Floor for quality observation. Raise to 3 during ramp-up.",            "visits_per_dp_week")
row = input_row(row, "RP working days per week",                     6, "days",    "Sunday off. If 7-day ops, increase but rotate RPs to keep 6/each.",     "rp_days_per_week")
row = input_row(row, "DPs observable per RP per day",                6, "DPs",     "3 on truck-follow + 3 independent. Hard cap given a 9-hr shift.",      "dps_per_rp_day")
row = input_row(row, "Headroom buffer (rain, illness, leave)",    1.10, "×",       "10% default. Set to 1.15 in monsoon months.",                          "buffer_factor", fmt="0.00")

row += 1

# ── Section: Derived programme structure ──────────────────────────────────
row = section_header(row, "2. Programme structure (derived)")
row = output_row(row, "Trucks required",
                 "=CEILING(meals_per_day/meals_per_truck,1)",
                 "trucks", "Fleet sizing — fed to truck-rotation logic.", "trucks_required")
row = output_row(row, "DPs required",
                 "=CEILING(meals_per_day/meals_per_dp,1)",
                 "DPs", "Implicit network size at this scale.", "dps_required")
row = output_row(row, "DPs per truck (avg)",
                 "=IFERROR(dps_required/trucks_required,0)",
                 "DPs/truck", "FILO-loading caps this at ~3. Above 3 → add trucks.", "dps_per_truck", fmt="0.0")
row = output_row(row, "Total DP-visits required per week",
                 "=dps_required*visits_per_dp_week",
                 "visits/wk", "What the RP team must collectively deliver.", "dp_visits_required")
row = output_row(row, "DP-visits one RP can deliver per week",
                 "=dps_per_rp_day*rp_days_per_week",
                 "visits/wk/RP", "Hard ceiling per RP given the SOP timeline.", "dp_visits_per_rp")

row += 1

# ── Section: RP team (the answer) ─────────────────────────────────────────
row = section_header(row, "3. RP team required")
row = output_row(row, "Kitchen-RPs needed (one per kitchen)",
                 "=num_kitchens",
                 "RPs", "Pre-dawn audit is single-kitchen; not parallelisable per RP.", "kitchen_rps")
row = output_row(row, "Field-RPs needed (by DP-visit load)",
                 "=CEILING(dp_visits_required/dp_visits_per_rp,1)",
                 "RPs", "Each RP doubles as field-RP after kitchen-shift ends.", "field_rps")
row = output_row(row, "Total RPs before buffer",
                 "=MAX(kitchen_rps,field_rps)",
                 "RPs", "Same person fills kitchen + field shifts.", "rps_pre_buffer")
row = output_row(row, "Total RPs required",
                 "=CEILING(rps_pre_buffer*buffer_factor,1)",
                 "RPs", "Headline number — staff to this.",
                 "total_rps", headline=True)
row = output_row(row, "RP : DP ratio",
                 "=IFERROR(total_rps/dps_required,0)",
                 "RP per DP", "Sanity gauge. Launch should sit ~0.2.", "rp_to_dp_ratio", fmt="0.00")
row = output_row(row, "RP : Truck ratio",
                 "=IFERROR(total_rps/trucks_required,0)",
                 "RP per truck", "Sanity gauge. Below 0.5 means truck rotation slips.", "rp_to_truck_ratio", fmt="0.00")

row += 1

# ── Section: Cost (inputs) ────────────────────────────────────────────────
row = section_header(row, "4. RP unit costs (inputs)")
row = input_row(row, "RP salary per month",                       50000, "₹/mo", "Mid-level field officer; Bangalore.",                "rp_salary_pm", fmt=FMT_INR)
row = input_row(row, "Two-wheeler allowance per month",            3500, "₹/mo", "Fuel + maintenance flat. Receipts filed quarterly.", "rp_2w_pm",    fmt=FMT_INR)
row = input_row(row, "Night cab per trip (01:30 home → kitchen)",   350, "₹",     "Programme-paid. Hard safety floor.",                 "rp_night_cab", fmt=FMT_INR)
row = input_row(row, "Buffer cab budget per RP per month",         3000, "₹/mo", "Monsoon, illness, late-night returns at ZL discretion.","rp_buffer_cab",fmt=FMT_INR)
row = input_row(row, "Working days per RP per month",                26, "days", "Excludes weekly off.",                                "days_per_month")

row += 1

# ── Section: Cost (derived) ───────────────────────────────────────────────
row = section_header(row, "5. Monthly cost (derived)")
row = output_row(row, "Salary",
                 "=total_rps*rp_salary_pm",
                 "₹/mo", "", "cost_salary", fmt=FMT_INR)
row = output_row(row, "Two-wheeler allowance",
                 "=total_rps*rp_2w_pm",
                 "₹/mo", "", "cost_2w", fmt=FMT_INR)
row = output_row(row, "Night cabs",
                 "=total_rps*days_per_month*rp_night_cab",
                 "₹/mo", "Every working day.", "cost_night_cab", fmt=FMT_INR)
row = output_row(row, "Buffer cabs",
                 "=total_rps*rp_buffer_cab",
                 "₹/mo", "", "cost_buffer_cab", fmt=FMT_INR)
row = output_row(row, "Total monthly RP cost",
                 "=cost_salary+cost_2w+cost_night_cab+cost_buffer_cab",
                 "₹/mo", "Add this line to the grant budget.",
                 "cost_total_monthly", fmt=FMT_INR, headline=True)
row = output_row(row, "RP cost per meal",
                 "=IFERROR(cost_total_monthly/(meals_per_day*days_per_month),0)",
                 "₹/meal", "Useful to compare against Wipro ₹20 / Ramani ₹29.40 production cost.",
                 "cost_per_meal", fmt=FMT_INR_DEC)
row = output_row(row, "RP cost as % of total opex",
                 "=IFERROR(cost_per_meal/36.27,0)",
                 "%", "Denominator = Sampark all-in ₹36.27/meal. Rough only.",
                 "cost_share", fmt=FMT_PCT)

row += 1
c = ws.cell(row=row, column=2,
            value="When inputs change, save the workbook. All blue/amber cells recalculate.")
c.font = HINT
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)


# ===========================================================================
# Sheet 2 — Scenarios
# ===========================================================================
ws2 = wb.create_sheet("Scenarios")

ws2.column_dimensions["A"].width = 4
ws2.column_dimensions["B"].width = 42
for col in range(3, 7):
    ws2.column_dimensions[get_column_letter(col)].width = 16

ws2.cell(row=1, column=2, value="Pre-loaded scenarios").font = TITLE
ws2.merge_cells(start_row=1, start_column=2, end_row=1, end_column=6)
c = ws2.cell(row=2, column=2,
             value="These are static reference points. Add columns to the right for new scenarios; do not overwrite columns C–E.")
c.font = HINT
ws2.merge_cells(start_row=2, start_column=2, end_row=2, end_column=6)

scenarios = [
    # (label, meals, kitchens, meals_per_truck, meals_per_dp, visits_per_dp_wk, rp_days, dps_per_rp_day, buffer)
    ("Launch (Sampark)",        1500,  1,  750, 300, 2, 6, 6, 1.10),
    ("Mid-scale",               5000,  1,  750, 300, 2, 6, 6, 1.10),
    ("Full Ramani (long-term)", 20000, 2, 1000, 350, 2, 6, 6, 1.15),
]

# Row layout: 4 = labels row
HEADER_ROW = 4
ws2.cell(row=HEADER_ROW, column=2, value="Driver").font = H2
ws2.cell(row=HEADER_ROW, column=2).fill = SECTION_FILL
for i, sc in enumerate(scenarios):
    cell = ws2.cell(row=HEADER_ROW, column=3 + i, value=sc[0])
    cell.font = H2
    cell.fill = SECTION_FILL
    cell.alignment = CENTER

def s_row(r, label, values, fmt=FMT_INT, fill=None, font=BODY):
    style_cell(ws2.cell(row=r, column=2, value=label), font=font, align=LEFT)
    for i, v in enumerate(values):
        c = ws2.cell(row=r, column=3 + i, value=v)
        style_cell(c, font=INPUT_FONT if fill is INPUT_FILL else font,
                   fill=fill, align=RIGHT, fmt=fmt)
    return r + 1

# Inputs block
r = HEADER_ROW + 1
ws2.cell(row=r, column=2, value="Inputs").font = H2
ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
r += 1
r = s_row(r, "Meals/day",            [s[1] for s in scenarios], fill=INPUT_FILL)
r = s_row(r, "Kitchens",             [s[2] for s in scenarios], fill=INPUT_FILL)
r = s_row(r, "Meals/truck/day",      [s[3] for s in scenarios], fill=INPUT_FILL)
r = s_row(r, "Meals/DP/day",         [s[4] for s in scenarios], fill=INPUT_FILL)
r = s_row(r, "Visits/DP/week",       [s[5] for s in scenarios], fill=INPUT_FILL)
r = s_row(r, "RP days/week",         [s[6] for s in scenarios], fill=INPUT_FILL)
r = s_row(r, "DPs/RP/day",           [s[7] for s in scenarios], fill=INPUT_FILL)
r = s_row(r, "Buffer factor",        [s[8] for s in scenarios], fmt="0.00", fill=INPUT_FILL)

r += 1

# Derived
ws2.cell(row=r, column=2, value="Derived").font = H2
ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
r += 1

def f_row(r, label, formula_per_col, fmt=FMT_INT, headline=False):
    style_cell(ws2.cell(row=r, column=2, value=label),
               font=TOTAL_FONT if headline else BODY, align=LEFT)
    for i in range(len(scenarios)):
        col = get_column_letter(3 + i)
        f = formula_per_col(col)
        c = ws2.cell(row=r, column=3 + i, value=f)
        style_cell(
            c,
            font=TOTAL_FONT if headline else OUTPUT_FONT,
            fill=TOTAL_FILL if headline else OUTPUT_FILL,
            align=RIGHT, fmt=fmt,
        )
    return r + 1

# Rows of inputs we need to reference (in declaration order):
ROW_MEALS    = HEADER_ROW + 2
ROW_KITCH    = HEADER_ROW + 3
ROW_PER_TRK  = HEADER_ROW + 4
ROW_PER_DP   = HEADER_ROW + 5
ROW_VIS_WK   = HEADER_ROW + 6
ROW_RP_DAYS  = HEADER_ROW + 7
ROW_DPS_DAY  = HEADER_ROW + 8
ROW_BUFFER   = HEADER_ROW + 9

r = f_row(r, "Trucks required",
          lambda c: f"=CEILING({c}{ROW_MEALS}/{c}{ROW_PER_TRK},1)")
r = f_row(r, "DPs required",
          lambda c: f"=CEILING({c}{ROW_MEALS}/{c}{ROW_PER_DP},1)")
r = f_row(r, "DPs per truck (avg)",
          lambda c: (f"=IFERROR(CEILING({c}{ROW_MEALS}/{c}{ROW_PER_DP},1)"
                     f"/CEILING({c}{ROW_MEALS}/{c}{ROW_PER_TRK},1),0)"),
          fmt="0.0")
r = f_row(r, "DP-visits required / week",
          lambda c: f"=CEILING({c}{ROW_MEALS}/{c}{ROW_PER_DP},1)*{c}{ROW_VIS_WK}")
r = f_row(r, "DP-visits / RP / week",
          lambda c: f"={c}{ROW_DPS_DAY}*{c}{ROW_RP_DAYS}")
r = f_row(r, "Kitchen-RPs",
          lambda c: f"={c}{ROW_KITCH}")
r = f_row(r, "Field-RPs (by DP load)",
          lambda c: (f"=CEILING("
                     f"(CEILING({c}{ROW_MEALS}/{c}{ROW_PER_DP},1)*{c}{ROW_VIS_WK})"
                     f"/({c}{ROW_DPS_DAY}*{c}{ROW_RP_DAYS}),1)"))
r = f_row(r, "RPs before buffer",
          lambda c: (f"=MAX({c}{ROW_KITCH},CEILING("
                     f"(CEILING({c}{ROW_MEALS}/{c}{ROW_PER_DP},1)*{c}{ROW_VIS_WK})"
                     f"/({c}{ROW_DPS_DAY}*{c}{ROW_RP_DAYS}),1))"))
r = f_row(r, "Total RPs required",
          lambda c: (f"=CEILING(MAX({c}{ROW_KITCH},CEILING("
                     f"(CEILING({c}{ROW_MEALS}/{c}{ROW_PER_DP},1)*{c}{ROW_VIS_WK})"
                     f"/({c}{ROW_DPS_DAY}*{c}{ROW_RP_DAYS}),1))*{c}{ROW_BUFFER},1)"),
          headline=True)


# ===========================================================================
# Sheet 3 — Staffing model (all roles, headcount, hours, monthly cost)
# Sources: lib/budget-costs.ts (food.*) + lib/line-template-seeds.ts
# ===========================================================================
ws_staff = wb.create_sheet("Staffing model")

ws_staff.column_dimensions["A"].width = 3
ws_staff.column_dimensions["B"].width = 32   # Role
ws_staff.column_dimensions["C"].width = 22   # Scaling basis
ws_staff.column_dimensions["D"].width = 10   # Ratio
ws_staff.column_dimensions["E"].width = 9    # Hrs/day
ws_staff.column_dimensions["F"].width = 10   # Headcount
ws_staff.column_dimensions["G"].width = 13   # Person-hrs/day
ws_staff.column_dimensions["H"].width = 14   # Monthly salary
ws_staff.column_dimensions["I"].width = 14   # Monthly cost
ws_staff.column_dimensions["J"].width = 28   # Cost key

r = 1
c = ws_staff.cell(row=r, column=2, value="Staffing model — all roles, at the scale set on the Calculator sheet")
c.font = TITLE
ws_staff.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)
r += 1
c = ws_staff.cell(row=r, column=2,
                  value="Edit mint-green cells (ratio / hrs-per-day / monthly salary) to override defaults. "
                        "Headcount and cost are derived from the Calculator sheet's meals_per_day, trucks_required, dps_required, total_rps.")
c.font = HINT
ws_staff.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)
r += 2

# ── linked inputs from Calculator ──
def section(r, title, span=10):
    c = ws_staff.cell(row=r, column=2, value=title)
    c.font = H1; c.fill = SECTION_FILL; c.alignment = LEFT
    for col in range(3, span + 1):
        ws_staff.cell(row=r, column=col).fill = SECTION_FILL
    ws_staff.merge_cells(start_row=r, start_column=2, end_row=r, end_column=span)
    return r + 1

r = section(r, "Inputs (auto-linked to Calculator sheet)")
def linked(r, label, name):
    style_cell(ws_staff.cell(row=r, column=2, value=label), font=BODY, align=LEFT)
    vc = ws_staff.cell(row=r, column=3, value=f"={name}")
    style_cell(vc, font=OUTPUT_FONT, fill=OUTPUT_FILL, align=RIGHT, fmt=FMT_INT)
    return r + 1
r = linked(r, "Meals per day",        "meals_per_day")
r = linked(r, "Kitchens",             "num_kitchens")
r = linked(r, "Trucks required",      "trucks_required")
r = linked(r, "DPs required",         "dps_required")
r = linked(r, "RPs required (from Calculator)", "total_rps")
r += 1

# ── headers ──
r = section(r, "Per-role staffing")
HEADERS = ["Role", "Scaling basis", "Ratio", "Hrs/day", "Headcount",
           "Person-hrs/day", "Monthly cost / person", "Monthly cost (total)", "Cost key"]
for i, h in enumerate(HEADERS):
    c = ws_staff.cell(row=r, column=2 + i, value=h)
    style_cell(c, font=H2, fill=SECTION_FILL, align=CENTER)
r += 1

# Roles: (label, scaling_kind, ratio, hrs_per_day, monthly_cost, cost_key)
# scaling_kind:
#   FIXED      → headcount = ratio (literal count)
#   PER_MEAL   → headcount = CEILING(meals_per_day / ratio, 1)
#   PER_TRUCK  → headcount = ratio * trucks_required
#   PER_DP     → headcount = ratio * dps_required
#   LINKED_RP  → headcount = total_rps  (ratio shown as text "—")
ROLES = [
    ("Programme Coordinator",       "FIXED",     1,      8,    65000, "food.programme_coordinator_salary"),
    ("Procurement Coordinator",     "FIXED",     1,     10,    50000, "food.procurement_coordinator_salary"),
    ("Delivery Coordinator",        "FIXED",     1,      9,    30000, "food.delivery_coordinator_salary"),
    ("Kitchen Manager",             "PER_MEAL",  10000,  9,    55000, "food.kitchen_manager_salary"),
    ("Warehouse Manager",           "PER_MEAL",  10000,  8,    40000, "food.warehouse_manager_salary"),
    ("Cook",                        "PER_MEAL",  1667,   4.5,  50000, "food.cook_salary"),
    ("Helper Cook",                 "PER_MEAL",  1112,   4.5,  25000, "food.helper_cook_salary"),
    ("Kitchen Loader",              "PER_MEAL",  1667,   3.0,  25000, "food.kitchen_loader_salary"),
    ("Chopping & Cleaning",         "PER_MEAL",  667,    8,    20000, "food.chopping_cleaning_salary"),
    ("Food Loader",                 "PER_MEAL",  667,    4,    18000, "food.food_loader_salary"),
    ("Housekeeping",                "PER_MEAL",  667,    8,    15000, "food.housekeeping_salary"),
    ("Truck (driver + vehicle, all-in)", "PER_TRUCK", 1, 5.5,  53100, "food.truck_cost_per_month"),
    ("DP Staff",                    "PER_DP",    2,      3,     6000, "food.dp_staff_remuneration_per_month"),
    ("Resource Person (RP)",        "LINKED_RP", None,   9,    50000, "rp_salary_pm (Calculator)"),
]

start_role_row = r
for (label, kind, ratio, hpd, cost_pp, key) in ROLES:
    style_cell(ws_staff.cell(row=r, column=2, value=label), font=BODY, align=LEFT)
    # basis label
    basis_label = {
        "FIXED":     "Fixed",
        "PER_MEAL":  "1 per X meals/day",
        "PER_TRUCK": "per truck",
        "PER_DP":    "per DP",
        "LINKED_RP": "RP calculator",
    }[kind]
    style_cell(ws_staff.cell(row=r, column=3, value=basis_label), font=HINT, align=LEFT)
    # ratio (editable input where applicable)
    if kind == "LINKED_RP":
        style_cell(ws_staff.cell(row=r, column=4, value="—"), font=HINT, align=CENTER)
    else:
        style_cell(ws_staff.cell(row=r, column=4, value=ratio),
                   font=INPUT_FONT, fill=INPUT_FILL, align=RIGHT,
                   fmt=FMT_INT if isinstance(ratio, int) and ratio >= 100 else "0")
    # hrs/day (editable)
    style_cell(ws_staff.cell(row=r, column=5, value=hpd),
               font=INPUT_FONT, fill=INPUT_FILL, align=RIGHT, fmt="0.0")
    # headcount formula
    D_ref = f"D{r}"
    if kind == "FIXED":
        hc_formula = f"={D_ref}"
    elif kind == "PER_MEAL":
        hc_formula = f"=CEILING(meals_per_day/{D_ref},1)"
    elif kind == "PER_TRUCK":
        hc_formula = f"={D_ref}*trucks_required"
    elif kind == "PER_DP":
        hc_formula = f"={D_ref}*dps_required"
    elif kind == "LINKED_RP":
        hc_formula = "=total_rps"
    style_cell(ws_staff.cell(row=r, column=6, value=hc_formula),
               font=OUTPUT_FONT, fill=OUTPUT_FILL, align=RIGHT, fmt=FMT_INT)
    # person-hours/day = headcount × hrs
    style_cell(ws_staff.cell(row=r, column=7, value=f"=F{r}*E{r}"),
               font=OUTPUT_FONT, fill=OUTPUT_FILL, align=RIGHT, fmt="0.0")
    # monthly cost / person (editable)
    style_cell(ws_staff.cell(row=r, column=8, value=cost_pp),
               font=INPUT_FONT, fill=INPUT_FILL, align=RIGHT, fmt=FMT_INR)
    # monthly cost total
    style_cell(ws_staff.cell(row=r, column=9, value=f"=F{r}*H{r}"),
               font=OUTPUT_FONT, fill=OUTPUT_FILL, align=RIGHT, fmt=FMT_INR)
    # cost key
    style_cell(ws_staff.cell(row=r, column=10, value=key), font=HINT, align=LEFT)
    r += 1

end_role_row = r - 1

# Totals row
style_cell(ws_staff.cell(row=r, column=2, value="TOTAL — staffing"),
           font=TOTAL_FONT, fill=TOTAL_FILL, align=LEFT)
for col in (3, 4, 5):
    style_cell(ws_staff.cell(row=r, column=col, value=""), font=BODY, fill=TOTAL_FILL)
style_cell(ws_staff.cell(row=r, column=6, value=f"=SUM(F{start_role_row}:F{end_role_row})"),
           font=TOTAL_FONT, fill=TOTAL_FILL, align=RIGHT, fmt=FMT_INT)
style_cell(ws_staff.cell(row=r, column=7, value=f"=SUM(G{start_role_row}:G{end_role_row})"),
           font=TOTAL_FONT, fill=TOTAL_FILL, align=RIGHT, fmt="0")
style_cell(ws_staff.cell(row=r, column=8, value=""), font=BODY, fill=TOTAL_FILL)
style_cell(ws_staff.cell(row=r, column=9, value=f"=SUM(I{start_role_row}:I{end_role_row})"),
           font=TOTAL_FONT, fill=TOTAL_FILL, align=RIGHT, fmt=FMT_INR)
style_cell(ws_staff.cell(row=r, column=10, value=""), font=HINT, fill=TOTAL_FILL)

# Save defined names for the totals so the headline can pick them up.
wb.defined_names["staff_total_headcount"] = DefinedName(name="staff_total_headcount", attr_text=f"'Staffing model'!$F${r}")
wb.defined_names["staff_total_personhrs"] = DefinedName(name="staff_total_personhrs", attr_text=f"'Staffing model'!$G${r}")
wb.defined_names["staff_total_monthly"]   = DefinedName(name="staff_total_monthly",   attr_text=f"'Staffing model'!$I${r}")
totals_row = r
r += 2

# ── kitchen utilities ─────────────────────────────────────────────────────
r = section(r, "Kitchen utility costs (per kitchen, monthly)")
UTILS = [
    ("Electricity",        75000, "food.electricity_per_month"),
    ("Water bill",         40000, "food.water_bill_per_month"),
    ("Cleaning",           60000, "food.cleaning_per_month"),
    ("Gas",                85000, "food.gas_per_month"),
    ("Maintenance & misc", 50000, "food.maintenance_per_month"),
]
util_start = r
for label, cost, key in UTILS:
    style_cell(ws_staff.cell(row=r, column=2, value=label), font=BODY, align=LEFT)
    style_cell(ws_staff.cell(row=r, column=3, value="per kitchen"), font=HINT, align=LEFT)
    style_cell(ws_staff.cell(row=r, column=8, value=cost),
               font=INPUT_FONT, fill=INPUT_FILL, align=RIGHT, fmt=FMT_INR)
    style_cell(ws_staff.cell(row=r, column=9, value=f"=num_kitchens*H{r}"),
               font=OUTPUT_FONT, fill=OUTPUT_FILL, align=RIGHT, fmt=FMT_INR)
    style_cell(ws_staff.cell(row=r, column=10, value=key), font=HINT, align=LEFT)
    r += 1
util_end = r - 1
# Subtotal
style_cell(ws_staff.cell(row=r, column=2, value="Subtotal — kitchen utilities"),
           font=TOTAL_FONT, fill=TOTAL_FILL, align=LEFT)
style_cell(ws_staff.cell(row=r, column=9, value=f"=SUM(I{util_start}:I{util_end})"),
           font=TOTAL_FONT, fill=TOTAL_FILL, align=RIGHT, fmt=FMT_INR)
wb.defined_names["utilities_monthly"] = DefinedName(name="utilities_monthly", attr_text=f"'Staffing model'!$I${r}")
utilities_row = r
r += 2

# ── DP consumables ────────────────────────────────────────────────────────
r = section(r, "DP consumables (monthly, scales with DPs and meals)")
# (label, unit_cost, per_unit_basis, monthly_qty_or_formula, key)
CONSUM = [
    # paper plates scale with meals*days_per_month
    ("Paper plates",       1.30,  "₹/plate",  "=meals_per_day*days_per_month",     "food.paper_plate_cost", "₹1.30 × meals/day × days/month"),
    ("Dustbin covers",     10,    "₹/cover",  "=dps_required*50",                  "food.dustbin_covers_per_dp_per_month", "50 covers/DP/month"),
    ("Gloves",             5,     "₹/pair",   "=dps_required*100",                 "food.gloves_per_dp_per_month", "100 pairs/DP/month"),
    ("Head caps",          2,     "₹/cap",    "=dps_required*50",                  "food.head_caps_per_dp_per_month", "50 caps/DP/month"),
    ("Drinking water cans",30,    "₹/can",    "=dps_required*50",                  "food.drinking_water_cans_per_dp_per_month", "50 cans/DP/month"),
    ("Aprons (annualised)",500,   "₹/apron",  "=dps_required*2/12",                "food.aprons_per_dp_per_year", "2 aprons/DP/year ÷ 12"),
    ("Misc DP supplies",   2000,  "₹/DP/mo",  "=dps_required",                     "food.misc_per_dp_per_month", "Flat per DP"),
]
consum_start = r
for label, unit_cost, unit, qty_formula, key, note in CONSUM:
    style_cell(ws_staff.cell(row=r, column=2, value=label), font=BODY, align=LEFT)
    style_cell(ws_staff.cell(row=r, column=3, value=note), font=HINT, align=LEFT)
    style_cell(ws_staff.cell(row=r, column=4, value=unit_cost),
               font=INPUT_FONT, fill=INPUT_FILL, align=RIGHT, fmt=FMT_INR_DEC)
    style_cell(ws_staff.cell(row=r, column=5, value=unit), font=HINT, align=CENTER)
    # quantity / month (computed)
    style_cell(ws_staff.cell(row=r, column=6, value=qty_formula),
               font=OUTPUT_FONT, fill=OUTPUT_FILL, align=RIGHT, fmt=FMT_INT)
    # monthly cost = unit cost × qty
    style_cell(ws_staff.cell(row=r, column=9, value=f"=D{r}*F{r}"),
               font=OUTPUT_FONT, fill=OUTPUT_FILL, align=RIGHT, fmt=FMT_INR)
    style_cell(ws_staff.cell(row=r, column=10, value=key), font=HINT, align=LEFT)
    r += 1
consum_end = r - 1
# Subtotal
style_cell(ws_staff.cell(row=r, column=2, value="Subtotal — DP consumables"),
           font=TOTAL_FONT, fill=TOTAL_FILL, align=LEFT)
style_cell(ws_staff.cell(row=r, column=9, value=f"=SUM(I{consum_start}:I{consum_end})"),
           font=TOTAL_FONT, fill=TOTAL_FILL, align=RIGHT, fmt=FMT_INR)
wb.defined_names["consumables_monthly"] = DefinedName(name="consumables_monthly", attr_text=f"'Staffing model'!$I${r}")
consumables_row = r
r += 2

# ── Food cost (the big one) ───────────────────────────────────────────────
r = section(r, "Food (raw material / vendor production)")
style_cell(ws_staff.cell(row=r, column=2, value="Food cost per meal"), font=BODY, align=LEFT)
style_cell(ws_staff.cell(row=r, column=3, value="Sampark/Ramani ₹29.40; CFAR/Wipro ₹20.00; veg ref ₹21.91"),
           font=HINT, align=LEFT)
style_cell(ws_staff.cell(row=r, column=4, value=29.40),
           font=INPUT_FONT, fill=INPUT_FILL, align=RIGHT, fmt=FMT_INR_DEC)
style_cell(ws_staff.cell(row=r, column=5, value="₹/meal"), font=HINT, align=CENTER)
style_cell(ws_staff.cell(row=r, column=9,
                         value=f"=D{r}*meals_per_day*days_per_month"),
           font=OUTPUT_FONT, fill=OUTPUT_FILL, align=RIGHT, fmt=FMT_INR)
style_cell(ws_staff.cell(row=r, column=10, value="food.cost_per_meal"), font=HINT, align=LEFT)
wb.defined_names["food_monthly"] = DefinedName(name="food_monthly", attr_text=f"'Staffing model'!$I${r}")
food_row = r
r += 2

# ── Headline grand total ──────────────────────────────────────────────────
style_cell(ws_staff.cell(row=r, column=2, value="GRAND TOTAL — monthly programme cost"),
           font=TOTAL_FONT, fill=TOTAL_FILL, align=LEFT)
ws_staff.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
style_cell(ws_staff.cell(row=r, column=9,
                         value="=staff_total_monthly+utilities_monthly+consumables_monthly+food_monthly"),
           font=TOTAL_FONT, fill=TOTAL_FILL, align=RIGHT, fmt=FMT_INR)
style_cell(ws_staff.cell(row=r, column=10, value=""), font=HINT, fill=TOTAL_FILL)
grand_row = r
r += 1
style_cell(ws_staff.cell(row=r, column=2, value="Cost per meal (all-in, monthly cost ÷ monthly meals)"),
           font=BODY, align=LEFT)
ws_staff.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
style_cell(ws_staff.cell(row=r, column=9,
                         value=f"=IFERROR(I{grand_row}/(meals_per_day*days_per_month),0)"),
           font=OUTPUT_FONT, fill=OUTPUT_FILL, align=RIGHT, fmt=FMT_INR_DEC)


# ===========================================================================
# Sheet 4 — Cost registry (mirror of lib/budget-costs.ts food.*)
# ===========================================================================
ws_reg = wb.create_sheet("Cost registry")

ws_reg.column_dimensions["A"].width = 3
ws_reg.column_dimensions["B"].width = 46  # cost key
ws_reg.column_dimensions["C"].width = 42  # description
ws_reg.column_dimensions["D"].width = 14  # unit cost
ws_reg.column_dimensions["E"].width = 22  # unit
ws_reg.column_dimensions["F"].width = 55  # notes

r = 1
c = ws_reg.cell(row=r, column=2, value="Cost registry — food.* from lib/budget-costs.ts")
c.font = TITLE
ws_reg.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
r += 1
c = ws_reg.cell(row=r, column=2,
                value="This is a mirror, not the source. When any value changes in lib/budget-costs.ts, "
                      "update this sheet AND the corresponding default on the Staffing model sheet.")
c.font = HINT
ws_reg.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
r += 2

# Headers
for i, h in enumerate(["Cost key", "Description", "Unit cost", "Unit", "Notes"]):
    c = ws_reg.cell(row=r, column=2 + i, value=h)
    style_cell(c, font=H2, fill=SECTION_FILL, align=CENTER)
r += 1

REGISTRY = [
    # (key, description, cost, unit, notes, section)
    # ── Food cost
    ("food.cost_per_meal",                   "Vegetarian menu reference",       21.91, "₹/meal",                "Weekly avg ₹153.36 ÷ 7 days", "FOOD"),
    # ── Coordinator salaries
    ("food.programme_coordinator_salary",    "Programme Coordinator",           65000, "₹/month",               "", "SALARY"),
    ("food.procurement_coordinator_salary",  "Procurement Coordinator",         50000, "₹/month",               "", "SALARY"),
    ("food.delivery_coordinator_salary",     "Delivery Coordinator",            30000, "₹/month",               "", "SALARY"),
    # ── Kitchen salaries
    ("food.kitchen_manager_salary",          "Kitchen Manager",                 55000, "₹/month",               "", "SALARY"),
    ("food.warehouse_manager_salary",        "Warehouse Manager",               40000, "₹/month",               "", "SALARY"),
    ("food.cook_salary",                     "Cook",                            50000, "₹/month",               "", "SALARY"),
    ("food.helper_cook_salary",              "Helper Cook",                     25000, "₹/month",               "", "SALARY"),
    ("food.kitchen_loader_salary",           "Kitchen Loader",                  25000, "₹/month",               "", "SALARY"),
    ("food.chopping_cleaning_salary",        "Chopping & Cleaning",             20000, "₹/month",               "", "SALARY"),
    ("food.food_loader_salary",              "Food Loader",                     18000, "₹/month",               "", "SALARY"),
    ("food.housekeeping_salary",             "Housekeeping",                    15000, "₹/month",               "", "SALARY"),
    # ── Kitchen ratios
    ("food.meals_per_kitchen_manager",       "Meals/day per Kitchen Manager",   10000, "meals/day per staff",   "1 Kitchen Mgr per 10K meals/day", "RATIO"),
    ("food.meals_per_warehouse_manager",     "Meals/day per Warehouse Manager", 10000, "meals/day per staff",   "1 Warehouse Mgr per 10K", "RATIO"),
    ("food.meals_per_cook",                  "Meals/day per Cook",              1667,  "meals/day per staff",   "6 Cooks per 10K", "RATIO"),
    ("food.meals_per_helper_cook",           "Meals/day per Helper Cook",       1112,  "meals/day per staff",   "9 Helper Cooks per 10K", "RATIO"),
    ("food.meals_per_kitchen_loader",        "Meals/day per Kitchen Loader",    1667,  "meals/day per staff",   "6 Loaders per 10K", "RATIO"),
    ("food.meals_per_chopping_cleaning",     "Meals/day per Chopping & Cleaning", 667, "meals/day per staff",   "15 per 10K", "RATIO"),
    ("food.meals_per_food_loader",           "Meals/day per Food Loader",       667,   "meals/day per staff",   "15 per 10K", "RATIO"),
    ("food.meals_per_housekeeping",          "Meals/day per Housekeeping",      667,   "meals/day per staff",   "15 per 10K", "RATIO"),
    # ── DP staff
    ("food.dp_staff_per_dp",                 "DP Staff per DP",                 2,     "staff per DP",          "", "RATIO"),
    ("food.dp_staff_remuneration_per_month", "DP Staff remuneration",           6000,  "₹/staff/month",         "Part-time, ~3 hrs/day", "SALARY"),
    # ── Kitchen utilities
    ("food.electricity_per_month",           "Electricity",                     75000, "₹/month",               "", "UTILITY"),
    ("food.water_bill_per_month",            "Water bill",                      40000, "₹/month",               "", "UTILITY"),
    ("food.cleaning_per_month",              "Cleaning (kitchen)",              60000, "₹/month",               "", "UTILITY"),
    ("food.gas_per_month",                   "Gas",                             85000, "₹/month",               "", "UTILITY"),
    ("food.maintenance_per_month",           "Maintenance & misc (kitchen)",    50000, "₹/month",               "", "UTILITY"),
    # ── Transport
    ("food.truck_cost_per_month",            "Truck (driver + fuel + maint + rental)", 53100, "₹/truck/month",  "JustDelivery all-in", "TRANSPORT"),
    # ── DP consumables
    ("food.paper_plate_cost",                "Paper plate",                     1.30,  "₹/plate",               "", "CONSUMABLE"),
    ("food.dustbin_cover_cost",              "Dustbin cover",                   10,    "₹/cover",               "", "CONSUMABLE"),
    ("food.dustbin_covers_per_dp_per_month", "Dustbin covers per DP per month", 50,    "covers/DP/month",       "", "CONSUMABLE"),
    ("food.gloves_cost",                     "Gloves",                          5,     "₹/pair",                "", "CONSUMABLE"),
    ("food.gloves_per_dp_per_month",         "Gloves per DP per month",         100,   "pairs/DP/month",        "", "CONSUMABLE"),
    ("food.head_cap_cost",                   "Head cap",                        2,     "₹/cap",                 "", "CONSUMABLE"),
    ("food.head_caps_per_dp_per_month",      "Head caps per DP per month",      50,    "caps/DP/month",         "", "CONSUMABLE"),
    ("food.drinking_water_can_cost",         "Drinking water can",              30,    "₹/can",                 "", "CONSUMABLE"),
    ("food.drinking_water_cans_per_dp_per_month", "Drinking water cans per DP per month", 50, "cans/DP/month",  "", "CONSUMABLE"),
    ("food.apron_cost",                      "Apron",                           500,   "₹/apron",               "", "CONSUMABLE"),
    ("food.aprons_per_dp_per_year",          "Aprons per DP per year",          2,     "aprons/DP/year",        "", "CONSUMABLE"),
    ("food.misc_per_dp_per_month",           "Misc DP supplies",                2000,  "₹/DP/month",            "", "CONSUMABLE"),
    # ── DP capex
    ("food.foldable_table_per_dp",           "Foldable table",                  6000,  "₹/DP",                  "One-time", "CAPEX"),
    ("food.canopy_tent_per_dp",              "Canopy tent",                     8000,  "₹/DP",                  "One-time", "CAPEX"),
    ("food.standee_umbrella_per_dp",         "Standee umbrella",                2000,  "₹/DP",                  "One-time", "CAPEX"),
    ("food.water_container_cost",            "Water container",                 250,   "₹/container",           "2 per DP", "CAPEX"),
    ("food.water_containers_per_dp",         "Water containers per DP",         2,     "containers/DP",         "", "CAPEX"),
    ("food.serving_kit_per_kitchen",         "Serving kit (per kitchen)",       85000, "₹/kitchen",             "Casseroles + vessels + spoons + chimta", "CAPEX"),
    ("food.kitchen_equipment_one_time",      "Kitchen equipment (in-house only)", 0,   "₹/kitchen",             "Set to ₹75.39L for in-house; 0 vendor-procured", "CAPEX"),
]

SECTION_COLORS = {
    "FOOD":       "FFFBEB",
    "SALARY":     "F0FDF4",
    "RATIO":      "EFF6FF",
    "UTILITY":    "F5F3FF",
    "TRANSPORT":  "FFF1F2",
    "CONSUMABLE": "F9FAFB",
    "CAPEX":      "FDF2F8",
}
current_section = None
for (key, desc, cost, unit, notes, sect) in REGISTRY:
    if sect != current_section:
        # section divider
        section_labels = {
            "FOOD":       "Food cost",
            "SALARY":     "Salaries",
            "RATIO":      "Staffing ratios",
            "UTILITY":    "Kitchen utilities",
            "TRANSPORT":  "Transport",
            "CONSUMABLE": "DP consumables",
            "CAPEX":      "DP capex / one-time",
        }
        c = ws_reg.cell(row=r, column=2, value=section_labels[sect])
        c.font = H2
        c.fill = PatternFill("solid", fgColor=SECTION_COLORS[sect])
        c.alignment = LEFT
        for col in range(3, 7):
            ws_reg.cell(row=r, column=col).fill = PatternFill("solid", fgColor=SECTION_COLORS[sect])
        ws_reg.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        r += 1
        current_section = sect
    style_cell(ws_reg.cell(row=r, column=2, value=key), font=BODY, align=LEFT)
    style_cell(ws_reg.cell(row=r, column=3, value=desc), font=BODY, align=LEFT)
    # cost format depends on magnitude
    cost_fmt = FMT_INR if (isinstance(cost, int) or cost >= 100) else FMT_INR_DEC
    style_cell(ws_reg.cell(row=r, column=4, value=cost),
               font=BODY, align=RIGHT, fmt=cost_fmt)
    style_cell(ws_reg.cell(row=r, column=5, value=unit), font=HINT, align=CENTER)
    style_cell(ws_reg.cell(row=r, column=6, value=notes), font=HINT, align=LEFT)
    r += 1


# ===========================================================================
# Sheet 5 — Assumptions
# ===========================================================================
ws3 = wb.create_sheet("Assumptions")
ws3.column_dimensions["A"].width = 4
ws3.column_dimensions["B"].width = 38
ws3.column_dimensions["C"].width = 92

ws3.cell(row=1, column=2, value="Assumptions & logic").font = TITLE
ws3.merge_cells(start_row=1, start_column=2, end_row=1, end_column=3)

ASSUMPTIONS = [
    ("Daily shift length",
     "RP works 02:00–11:00 = 9 hours. Kitchen morning + dispatch + truck-follow + independent DP rounds + reconciliation + daily wrap. See SOP §2."),
    ("DPs observable per RP per day",
     "Default 6: 3 DPs while shadowing one truck (truck serves ~3 DPs per route) + 3 more during independent rounds 07:30–09:30. Hard ceiling; do not raise without reducing dwell-time below 15 min/DP."),
    ("DP coverage rule",
     "Every DP must be observed during service at least twice per week. Below 2 we lose quality signal; above 3 the SOP becomes inspection theatre."),
    ("Truck rotation rule",
     "Every truck shadowed end-to-end at least twice per week, on a 6-day rolling roster."),
    ("Kitchen audit is single-kitchen per RP",
     "Pre-dawn audit + fill + FILO loading is a single physical location. Two kitchens cannot share one morning RP; staff one RP per kitchen as the floor."),
    ("Meals per truck",
     "TATA Ace × 6 containers × ~125 portions per container ≈ 750 meals/day. Edit if container count or portion size changes."),
    ("Meals per DP",
     "Default 300 (Sampark new DPs). CFAR legacy DPs run 150–200; new DPs 200. For a mixed-grant kitchen, use a weighted average."),
    ("Headroom buffer",
     "Multiplier on RP count to absorb monsoon, illness, training, leave. Default 1.10 (10%). Use 1.15 in June–September Bangalore monsoon months. Do not go below 1.05."),
    ("Working days per RP per month",
     "Default 26 (Sunday off, 4–5 Sundays per month). If the programme moves to 7-day ops, RPs rotate so each still works ≤ 6 days/week."),
    ("RP transport (cost model)",
     "Night cab (home → kitchen) every working day = 26 × ₹350. Two-wheeler allowance covers daytime fuel + maintenance. Buffer cab is at ZL discretion: monsoon, illness, late return. See SOP §1.3."),
    ("Staffing-model hours/day",
     "Hours/day per role are observed averages, not contracted hours. Most kitchen roles are paid monthly regardless of actual hours — these figures drive the person-hours rollup, not payroll."),
    ("Staffing-model ratios (source: lib/budget-costs.ts)",
     "Kitchen Manager 1:10K, Warehouse Manager 1:10K, Cook 1:1,667 (6 per 10K), Helper Cook 1:1,112 (9), Kitchen Loader 1:1,667 (6), Chopping & Cleaning 1:667 (15), Food Loader 1:667 (15), Housekeeping 1:667 (15), DP Staff 2 per DP, Truck Driver 1 per truck (rolled into ₹53,100/truck/mo). Coordinators (Programme/Procurement/Delivery) fixed at 1 each up to ~25K meals/day."),
    ("Truck cost is all-in",
     "₹53,100/truck/mo from food.truck_cost_per_month covers driver + fuel + maintenance + vehicle rental (JustDelivery). Do not separately add driver salary."),
    ("DP Staff are part-time",
     "₹6,000/staff/month at 2 staff per DP, ~3 hours/day. Bangalore rate; revise for other cities."),
    ("Food cost per meal — vendor vs in-house",
     "Defaults to Sampark/Ramani ₹29.40 (vendor). For CFAR/Wipro use ₹20.00. For in-house kitchen use the ₹21.91 reference plus add the kitchen_equipment_one_time capex line."),
    ("Cost registry sheet is a mirror",
     "When any value changes in lib/budget-costs.ts, update the Cost registry sheet AND the matching default on Staffing model. Cost registry has no formulas — it is reference only."),
    ("When to re-run",
     "Whenever any of: meals/day target changes, a new kitchen is added, fleet composition changes, DP network grows past +30%, RP shift hours change, OR any value in lib/budget-costs.ts changes. Also re-run before every quarterly grant review."),
]

r = 3
for label, body in ASSUMPTIONS:
    style_cell(ws3.cell(row=r, column=2, value=label), font=H2, align=LEFT)
    style_cell(ws3.cell(row=r, column=3, value=body), font=BODY, align=LEFT)
    ws3.row_dimensions[r].height = 38
    r += 1


# ── save ───────────────────────────────────────────────────────────────────
wb.save(OUT)
print(f"Wrote {OUT}")
